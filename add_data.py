# ============================================================
#  add_data.py  (v2 — 자동 ROI 예측 통합)
#  새 사진 추가 + Contact Point 위치 자동 예측
#
#  실행: python add_data.py
#
#  동작:
#  1. new_images/ 폴더의 사진을 data/images/ 로 복사
#  2. 등급 입력 (1~5)
#  3. roi_detector.pth 있으면 → Contact Point 자동 예측
#     없으면 → "python train_roi_detector.py 먼저 실행" 안내
#  4. 예측 결과를 roi_annotations.json 에 자동 저장
# ============================================================

import os, sys, json, shutil
import numpy as np
import pandas as pd
from PIL import Image, ImageDraw

# ── 설정 ────────────────────────────────────────────────────
IMAGE_DIR    = "data/images"
LABEL_CSV    = "data/labels.csv"
ROI_JSON     = "data/roi_annotations.json"
INBOX_DIR    = "new_images"
DETECTOR_PTH = "model/roi_detector.pth"

USE_EXCEL    = False
EXCEL_PATH   = "data/new_labels.xlsx"


# ── ROI 자동 예측 ────────────────────────────────────────────
def load_roi_detector():
    """roi_detector.pth 로드. 없으면 None 반환."""
    if not os.path.exists(DETECTOR_PTH):
        return None, None

    try:
        import torch
        import torch.nn as nn
        from torchvision import models, transforms

        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        ckpt   = torch.load(DETECTOR_PTH, map_location=device)

        class ContactPointDetector(nn.Module):
            def __init__(self):
                super().__init__()
                base = models.efficientnet_b0(weights=None)
                self.features  = base.features
                self.pool      = nn.AdaptiveAvgPool2d(1)
                self.regressor = nn.Sequential(
                    nn.Dropout(p=0.4),
                    nn.Linear(1280, 256),
                    nn.ReLU(inplace=True),
                    nn.Dropout(p=0.3),
                    nn.Linear(256, 4),
                    nn.Sigmoid(),
                )
            def forward(self, x):
                return self.regressor(self.pool(self.features(x)).flatten(1))

        model = ContactPointDetector().to(device)
        model.load_state_dict(ckpt["model_state_dict"])
        model.eval()

        img_size = ckpt.get("img_size", 224)
        val_iou  = ckpt.get("best_val_iou", 0)

        tf = transforms.Compose([
            transforms.Resize((img_size, img_size)),
            transforms.ToTensor(),
            transforms.Normalize([0.485,0.456,0.406],[0.229,0.224,0.225]),
        ])

        print(f"✅ ROI 자동 예측 모델 로드 (Val IoU: {val_iou:.3f})")
        return (model, device, tf), val_iou

    except Exception as e:
        print(f"⚠️ ROI 모델 로드 실패: {e}")
        return None, None


def predict_roi(model_bundle, pil_img):
    """PIL 이미지 → (x1, y1, x2, y2) 픽셀 좌표 예측"""
    model, device, tf = model_bundle
    W, H = pil_img.size
    import torch
    with torch.no_grad():
        pred = model(tf(pil_img).unsqueeze(0).to(device)).squeeze().cpu().numpy()
    x1 = max(0, int(pred[0] * W))
    y1 = max(0, int(pred[1] * H))
    x2 = min(W, int(pred[2] * W))
    y2 = min(H, int(pred[3] * H))
    return x1, y1, x2, y2


def show_preview_with_box(pil_img, x1, y1, x2, y2, save_path):
    """예측 박스를 이미지에 그려서 저장 (CMD 미리보기용)"""
    vis = pil_img.copy()
    d   = ImageDraw.Draw(vis)
    d.rectangle([x1, y1, x2, y2], outline="#ff6400", width=max(3, pil_img.width // 150))
    cs = max(10, pil_img.width // 60)
    for cx, cy in [(x1,y1),(x2,y1),(x1,y2),(x2,y2)]:
        d.rectangle([cx-cs, cy-cs, cx+cs, cy+cs], outline="#ff6400", width=4)
    vis.save(save_path)


# ── ROI JSON 로드/저장 ────────────────────────────────────────
def load_roi_map():
    if os.path.exists(ROI_JSON):
        with open(ROI_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_roi_map(roi_map):
    with open(ROI_JSON, "w", encoding="utf-8") as f:
        json.dump(roi_map, f, ensure_ascii=False, indent=2)


# ── 메인 폴더 모드 ────────────────────────────────────────────
def method1_folder():
    os.makedirs(INBOX_DIR, exist_ok=True)
    files = [f for f in os.listdir(INBOX_DIR)
             if f.lower().endswith((".jpg",".jpeg",".png"))]

    if not files:
        print(f"⚠️  '{INBOX_DIR}/' 폴더에 이미지가 없습니다.")
        print(f"   JPG/PNG 파일을 넣고 다시 실행하세요.")
        return

    print(f"\n✅ {len(files)}개 이미지 발견")

    # ROI 모델 로드
    model_bundle, val_iou = load_roi_detector()
    if model_bundle is None:
        print("\n💡 ROI 자동 예측 모델이 없습니다.")
        print("   먼저 실행하세요: python train_roi_detector.py")
        print("   (ROI 없이 사진/등급만 추가됩니다)\n")
    else:
        iou_level = "좋음" if val_iou >= 0.5 else "참고용 (수동 확인 권장)"
        print(f"   예측 품질: IoU {val_iou:.3f} ({iou_level})\n")

    # 기존 데이터 로드
    df      = pd.read_csv(LABEL_CSV) if os.path.exists(LABEL_CSV) \
              else pd.DataFrame(columns=["filename","grade"])
    roi_map = load_roi_map()
    existing = set(df["filename"].tolist())

    new_rows   = []
    roi_added  = 0
    preview_dir = os.path.join("new_images", "roi_preview")
    os.makedirs(preview_dir, exist_ok=True)

    print("각 이미지의 등급을 입력하세요 (1~5, 건너뛰려면 엔터)")
    print("=" * 55)

    for fname in sorted(files):
        if fname in existing:
            print(f"  {fname} → 이미 등록됨, 건너뜀")
            continue

        img_path = os.path.join(INBOX_DIR, fname)
        try:
            pil = Image.open(img_path).convert("RGB")
            W, H = pil.size
            print(f"\n  📷 {fname}  ({W}×{H}px)")
        except Exception:
            print(f"\n  📷 {fname}  (열기 실패, 건너뜀)")
            continue

        # 등급 입력
        while True:
            g_in = input("     등급 (1~5): ").strip()
            if g_in == "":
                print("     건너뜀")
                break
            if g_in in ["1","2","3","4","5"]:
                grade = int(g_in)
                new_rows.append({"filename": fname, "grade": grade})
                print(f"     ✅ Grade {grade} 등록")
                break
            print("     ❌ 1~5 중 하나 입력하세요")

        if not new_rows or new_rows[-1]["filename"] != fname:
            continue  # 건너뜀

        # ROI 자동 예측
        if model_bundle is not None and fname not in roi_map:
            x1, y1, x2, y2 = predict_roi(model_bundle, pil)
            print(f"     🎯 Contact Point 자동 예측: ({x1},{y1}) → ({x2},{y2})")

            # 미리보기 이미지 저장
            prev_path = os.path.join(preview_dir, f"roi_{fname}")
            show_preview_with_box(pil, x1, y1, x2, y2, prev_path)
            print(f"     🖼️  미리보기 저장: {prev_path}")

            # 사용자 확인
            confirm = input("     이 ROI 사용하시겠어요? (y/n, 기본 y): ").strip().lower()
            if confirm in ("", "y", "yes"):
                roi_map[fname] = {
                    "x1": x1, "y1": y1,
                    "x2": x2, "y2": y2,
                    "orig_w": W, "orig_h": H,
                    "auto_predicted": True,
                }
                save_roi_map(roi_map)
                roi_added += 1
                print(f"     ✅ ROI 자동 저장 완료")
            else:
                print(f"     ⚠️  ROI 미저장 → annotate.py 에서 수동 지정 가능")

    if not new_rows:
        print("\n추가된 이미지가 없습니다.")
        return

    # 이미지 복사
    copied = 0
    for row in new_rows:
        src = os.path.join(INBOX_DIR, row["filename"])
        dst = os.path.join(IMAGE_DIR,  row["filename"])
        if not os.path.exists(dst):
            shutil.copy2(src, dst)
            copied += 1

    # labels.csv 업데이트
    new_df = pd.DataFrame(new_rows)
    df     = pd.concat([df, new_df], ignore_index=True)
    df.to_csv(LABEL_CSV, index=False)

    print(f"\n{'='*55}")
    print(f"✅ 완료!")
    print(f"   이미지 {copied}개 복사 → {IMAGE_DIR}/")
    print(f"   labels.csv 업데이트 → 총 {len(df)}개")
    if roi_added:
        print(f"   ROI 자동 예측 저장 → {roi_added}개")
    print("\n📊 등급 분포:")
    print(df["grade"].value_counts().sort_index().to_string())
    print("\n📌 ROI가 잘못 예측된 사진은 annotate.py 에서 수정 가능합니다.")
    print("\n다음 단계: python train.py")


# ── 엑셀 모드 (기존 유지) ────────────────────────────────────
def method2_excel():
    import zipfile
    from openpyxl import load_workbook

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ 엑셀 파일을 찾을 수 없습니다: {EXCEL_PATH}")
        return

    print(f"📊 엑셀 파일 분석 중: {EXCEL_PATH}")
    wb   = load_workbook(EXCEL_PATH, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    grade_map = {}
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            if isinstance(val, (int, float)) and 1 <= val <= 5:
                grade_map[r_idx] = int(val)

    print(f"  등급 셀 {len(grade_map)}개 발견")

    os.makedirs(IMAGE_DIR, exist_ok=True)
    df = pd.read_csv(LABEL_CSV) if os.path.exists(LABEL_CSV) \
         else pd.DataFrame(columns=["filename","grade"])
    existing = set(df["filename"].tolist())

    # ROI 모델 로드
    model_bundle, val_iou = load_roi_detector()
    roi_map = load_roi_map()
    new_rows = []; roi_added = 0

    with zipfile.ZipFile(EXCEL_PATH, "r") as z:
        img_files = sorted([f for f in z.namelist()
                            if f.startswith("xl/media/")])
        anchors   = []
        for img_obj in ws._images:
            try:    anchors.append(img_obj.anchor._from.row)
            except: anchors.append(None)

        for i, img_file in enumerate(img_files):
            ext = img_file.rsplit(".", 1)[-1].lower()
            if ext not in ("jpg","jpeg","png","gif","bmp"): continue

            grade = None
            if i < len(anchors) and anchors[i] is not None:
                r = anchors[i]
                for delta in [0, 1,-1, 2,-2]:
                    if r + delta in grade_map:
                        grade = grade_map[r + delta]; break
            if grade is None:
                print(f"  ⚠️ {img_file} → 등급 매핑 실패, 건너뜀"); continue

            fname = f"new_{i+1:03d}_grade{grade}.jpg"
            dst   = os.path.join(IMAGE_DIR, fname)
            if fname in existing: continue

            import io as _io
            with z.open(img_file) as sf:
                pil = Image.open(_io.BytesIO(sf.read())).convert("RGB")
                pil.save(dst, "JPEG", quality=90)

            new_rows.append({"filename": fname, "grade": grade})

            # ROI 자동 예측
            if model_bundle and fname not in roi_map:
                W, H = pil.size
                x1, y1, x2, y2 = predict_roi(model_bundle, pil)
                roi_map[fname] = {
                    "x1": x1, "y1": y1, "x2": x2, "y2": y2,
                    "orig_w": W, "orig_h": H,
                    "auto_predicted": True,
                }
                roi_added += 1

    if new_rows:
        save_roi_map(roi_map)
        new_df = pd.DataFrame(new_rows)
        df     = pd.concat([df, new_df], ignore_index=True)
        df.to_csv(LABEL_CSV, index=False)
        print(f"\n✅ 완료! {len(new_rows)}개 이미지 추출/등록")
        if roi_added:
            print(f"   ROI 자동 예측 저장: {roi_added}개")
        print("\n📊 등급 분포:")
        print(df["grade"].value_counts().sort_index().to_string())
    else:
        print("⚠️ 새로 추가된 이미지가 없습니다.")


# ── Entry ────────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  CPC — 데이터 추가 도구 (v2: 자동 ROI 예측)")
    print("=" * 55)
    os.makedirs(IMAGE_DIR, exist_ok=True)

    if USE_EXCEL:
        print("\n📊 엑셀 모드\n")
        method2_excel()
    else:
        print(f"\n📂 폴더 모드: '{INBOX_DIR}/' 폴더에 사진을 넣고 실행\n")
        method1_folder()


if __name__ == "__main__":
    main()
