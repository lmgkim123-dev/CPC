# ============================================================
#  app.py  (v7 - ROI 직접 지정 추론 + annotate.py 연동)
#  Contact Point Corrosion 등급 분류 Streamlit 웹앱
#  실행: streamlit run app.py
# ============================================================

import os, io, datetime, base64
import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image, ImageDraw
import torch
import torch.nn as nn
from torchvision import models, transforms

# streamlit-drawable-canvas는 사용하지 않음 (Streamlit 최신버전 비호환)
# ROI는 슬라이더 + PIL 미리보기 방식으로 구현
CANVAS_AVAILABLE = False  # 하위 호환용 플래그 (미사용)

# Grad-CAM 모듈 (Contact Point 시각화)
try:
    from gradcam import analyze_contact_point, pil_to_bytes
    GRADCAM_AVAILABLE = True
except ImportError:
    GRADCAM_AVAILABLE = False

# ── 페이지 설정 ──────────────────────────────────
st.set_page_config(
    page_title="Contact Point Corrosion 등급 분류",
    page_icon="🔬",
    layout="wide",
)

MODEL_PATH  = "model/best_model.pth"
IMG_SIZE    = 224
NUM_CLASSES = 5
DEVICE      = torch.device("cuda" if torch.cuda.is_available() else "cpu")

GRADE_INFO = {
    1: {"label": "Grade 1 — 정상",           "color": "#2ecc71", "emoji": "🟢"},
    2: {"label": "Grade 2 — 경미한 스케일",   "color": "#f1c40f", "emoji": "🟡"},
    3: {"label": "Grade 3 — 중간 스케일",     "color": "#e67e22", "emoji": "🟠"},
    4: {"label": "Grade 4 — 심각한 스케일",   "color": "#e74c3c", "emoji": "🔴"},
    5: {"label": "Grade 5 — 매우 심각",       "color": "#8e44ad", "emoji": "🟣"},
}

# ── 클립보드 붙여넣기 JS ────────────────────────
CLIPBOARD_HTML = """
<style>
#paste-area {
    width: 100%;
    min-height: 160px;
    border: 2.5px dashed #3b82f6;
    border-radius: 12px;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    background: #f0f7ff;
    cursor: pointer;
    transition: all 0.2s;
    padding: 20px;
    box-sizing: border-box;
}
#paste-area:hover, #paste-area.drag-over {
    background: #dbeafe;
    border-color: #2563eb;
}
#paste-area p { color:#3b82f6;font-size:1rem;margin:6px 0;font-weight:500; }
#paste-area .icon { font-size:2.2rem;margin-bottom:4px; }
#preview-img {
    max-width:100%;max-height:300px;border-radius:10px;
    margin-top:10px;display:none;border:1px solid #cbd5e1;
}
#paste-status {
    margin-top:8px;color:#16a34a;font-weight:600;
    font-size:0.9rem;display:none;
}
</style>
<div id="paste-area" tabindex="0">
    <div class="icon">📋</div>
    <p>여기를 클릭 후 <b>Ctrl + V</b> 로 이미지 붙여넣기</p>
    <p style="font-size:0.82rem;color:#64748b;">캡쳐 도구로 복사한 이미지를 바로 붙여넣을 수 있어요</p>
</div>
<img id="preview-img" alt="붙여넣기 미리보기"/>
<div id="paste-status">✅ 이미지 붙여넣기 완료!</div>
<script>
const area=document.getElementById('paste-area');
const preview=document.getElementById('preview-img');
const status=document.getElementById('paste-status');
area.addEventListener('click',()=>area.focus());
area.addEventListener('dragover',(e)=>{e.preventDefault();area.classList.add('drag-over');});
area.addEventListener('dragleave',()=>area.classList.remove('drag-over'));
area.addEventListener('drop',(e)=>{
    e.preventDefault();area.classList.remove('drag-over');
    const file=e.dataTransfer.files[0];
    if(file&&file.type.startsWith('image/'))processFile(file);
});
document.addEventListener('paste',(e)=>{
    const items=e.clipboardData.items;
    for(let item of items){
        if(item.type.startsWith('image/')){processFile(item.getAsFile());break;}
    }
});
function processFile(file){
    const reader=new FileReader();
    reader.onload=(ev)=>{
        const dataUrl=ev.target.result;
        preview.src=dataUrl;preview.style.display='block';
        status.style.display='block';
        area.querySelector('p').style.display='none';
        const b64=dataUrl.split(',')[1];
        window.parent.postMessage({type:'streamlit:setComponentValue',value:b64},'*');
    };
    reader.readAsDataURL(file);
}
</script>
"""

# ── CBAM 모듈 (train.py 와 동일하게 정의) ─────────
class ChannelAttention(nn.Module):
    def __init__(self, in_channels, reduction=16):
        super().__init__()
        self.avg_pool = nn.AdaptiveAvgPool2d(1)
        self.max_pool = nn.AdaptiveMaxPool2d(1)
        r = max(1, in_channels // reduction)
        self.fc = nn.Sequential(
            nn.Linear(in_channels, r, bias=False),
            nn.ReLU(inplace=True),
            nn.Linear(r, in_channels, bias=False),
        )
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        b, c, _, _ = x.size()
        avg = self.fc(self.avg_pool(x).view(b, c))
        mx  = self.fc(self.max_pool(x).view(b, c))
        return x * self.sigmoid(avg + mx).view(b, c, 1, 1)


class SpatialAttention(nn.Module):
    def __init__(self, kernel_size=7):
        super().__init__()
        self.conv    = nn.Conv2d(2, 1, kernel_size, padding=kernel_size//2, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        avg = x.mean(dim=1, keepdim=True)
        mx, _ = x.max(dim=1, keepdim=True)
        return x * self.sigmoid(self.conv(torch.cat([avg, mx], dim=1)))


class CBAM(nn.Module):
    def __init__(self, in_channels, reduction=16, kernel_size=7):
        super().__init__()
        self.ch = ChannelAttention(in_channels, reduction)
        self.sp = SpatialAttention(kernel_size)

    def forward(self, x):
        return self.sp(self.ch(x))


class EfficientNetWithCBAM(nn.Module):
    def __init__(self, num_classes):
        super().__init__()
        base = models.efficientnet_b0(weights=None)
        self.features   = base.features
        self.cbam       = CBAM(1280, reduction=16, kernel_size=7)
        self.pool       = nn.AdaptiveAvgPool2d(1)
        self.classifier = nn.Sequential(
            nn.Dropout(p=0.4),
            nn.Linear(1280, 256),
            nn.ReLU(inplace=True),
            nn.Dropout(p=0.3),
            nn.Linear(256, num_classes),
        )

    def forward(self, x):
        feat = self.features(x)
        feat = self.cbam(feat)
        feat = self.pool(feat).flatten(1)
        return self.classifier(feat)


# ── 모델 로드 ─────────────────────────────────────
@st.cache_resource
def load_model():
    ckpt = torch.load(MODEL_PATH, map_location=DEVICE)
    use_cbam   = ckpt.get("use_cbam", False)
    num_classes = ckpt.get("num_classes", NUM_CLASSES)

    if use_cbam:
        model = EfficientNetWithCBAM(num_classes)
    else:
        # 구 버전 호환 (CBAM 없는 모델)
        base = models.efficientnet_b0(weights=None)
        in_f = base.classifier[1].in_features
        base.classifier = nn.Sequential(
            nn.Dropout(p=0.4),
            nn.Linear(in_f, 256),
            nn.ReLU(),
            nn.Dropout(p=0.3),
            nn.Linear(256, num_classes),
        )
        model = base

    model.load_state_dict(ckpt["model_state_dict"])
    model.to(DEVICE).eval()
    return model, use_cbam


def get_model_info():
    try:
        ckpt = torch.load(MODEL_PATH, map_location=DEVICE)
        return {
            "total_data"       : ckpt.get("total_data",   "?"),
            "total_train"      : ckpt.get("total_train",  "?"),
            "best_val_accuracy": ckpt.get("best_val_accuracy", 0),
            "grade_dist"       : ckpt.get("grade_dist",   {}),
            "use_cbam"         : ckpt.get("use_cbam", False),
        }
    except:
        return None


# ── 전처리 + 예측 ────────────────────────────────
def get_transform():
    return transforms.Compose([
        transforms.Resize((IMG_SIZE, IMG_SIZE)),
        transforms.ToTensor(),
        transforms.Normalize([0.485, 0.456, 0.406],
                             [0.229, 0.224, 0.225]),
    ])


def predict(model, pil_image):
    tf     = get_transform()
    tensor = tf(pil_image).unsqueeze(0).to(DEVICE)
    with torch.no_grad():
        probs = torch.softmax(model(tensor), dim=1).squeeze()
    grade      = int(probs.argmax().item()) + 1
    confidence = float(probs.max().item()) * 100
    all_probs  = {i+1: float(probs[i].item())*100 for i in range(NUM_CLASSES)}
    return grade, confidence, all_probs


# ── Excel + 사진 포함 내보내기 ───────────────────
def to_excel_with_images(df: pd.DataFrame, images: list) -> bytes:
    """
    df     : 결과 DataFrame (파일명, 등급, 설명, 신뢰도, 시각)
    images : PIL Image 리스트 (df 행과 동일 순서)
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = "부식등급결과"

    # ── 헤더 구성 (사진 열 추가) ──
    headers = ["사진"] + list(df.columns)
    ws.append(headers)

    # 헤더 스타일
    hdr_fill = PatternFill(start_color="1F4E79",
                           end_color="1F4E79", fill_type="solid")
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 행 높이 (사진 크기에 맞게)
    THUMB_W = 80   # px
    THUMB_H = 80   # px
    ROW_H   = 65   # pt (약 80px)

    grade_colors = {
        "1": "C6EFCE", "2": "FFEB9C",
        "3": "FFCC99", "4": "FFC7CE", "5": "E2AFEF"
    }

    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        ws.row_dimensions[row_idx].height = ROW_H

        # 사진 열(A) 비워두기 (이미지 삽입 예정)
        ws.cell(row=row_idx, column=1).value = ""

        # 데이터 열 채우기
        for col_idx, val in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            # 등급 열 색상
            if df.columns[col_idx - 2] == "예측 등급":
                gs = str(val)
                if gs in grade_colors:
                    cell.fill = PatternFill(
                        start_color=grade_colors[gs],
                        end_color=grade_colors[gs], fill_type="solid")

        # 썸네일 이미지 삽입
        if row_idx - 2 < len(images) and images[row_idx - 2] is not None:
            try:
                pil_img = images[row_idx - 2].copy()
                pil_img.thumbnail((THUMB_W, THUMB_H))
                img_buf = io.BytesIO()
                pil_img.save(img_buf, format="PNG")
                img_buf.seek(0)
                xl_img = XLImage(img_buf)
                xl_img.width  = THUMB_W
                xl_img.height = THUMB_H
                # 셀 위치 지정 (약간 여백)
                cell_addr = f"A{row_idx}"
                ws.add_image(xl_img, cell_addr)
            except Exception:
                pass

    # 열 너비 조정
    ws.column_dimensions["A"].width = 14   # 사진 열
    for col_idx in range(2, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ── UI 컴포넌트 ───────────────────────────────────
def render_header():
    st.markdown("""
    <div style='background:linear-gradient(135deg,#1a1a2e,#16213e);
                padding:28px 32px;border-radius:14px;margin-bottom:20px;'>
      <h1 style='color:#e2e8f0;margin:0;font-size:1.9rem;'>
        🔬 Contact Point Corrosion 등급 분류
      </h1>
      <p style='color:#94a3b8;margin:8px 0 0;'>
        EfficientNet-B0 + <b style='color:#60a5fa;'>CBAM Attention</b>
        기반 스케일 형성 자동 판정 시스템
      </p>
    </div>""", unsafe_allow_html=True)


def render_model_stats():
    info = get_model_info()
    if not info:
        return

    grade_colors = {
        1: ("#2ecc71","🟢"), 2: ("#f1c40f","🟡"),
        3: ("#e67e22","🟠"), 4: ("#e74c3c","🔴"), 5: ("#8e44ad","🟣")
    }
    acc = info["best_val_accuracy"] * 100
    acc_color = "#2ecc71" if acc >= 70 else "#e67e22" if acc >= 50 else "#e74c3c"
    cbam_badge = ("🧠 **CBAM Attention 적용** (Contact Point 자동 포착)"
                  if info.get("use_cbam") else "")

    if cbam_badge:
        st.success(cbam_badge)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📸 전체 학습 데이터",      f"{info['total_data']}장")
    c2.metric("🎓 학습에 사용된 데이터",  f"{info['total_train']}장")
    c3.metric("🎯 검증 정확도",           f"{acc:.1f}%")
    c4.metric("🏷️ 분류 등급",            "1 ~ 5 단계")

    if info["grade_dist"]:
        st.markdown("#### 📊 등급별 학습 데이터 현황")
        dist  = info["grade_dist"]
        total = sum(dist.values())
        cols  = st.columns(5)
        for i, col in enumerate(cols, 1):
            cnt   = dist.get(i, 0)
            color, emoji = grade_colors.get(i, ("#94a3b8","⬜"))
            pct   = cnt / total * 100 if total > 0 else 0
            col.markdown(f"""
            <div style='background:{color}18;border:1.5px solid {color};
                        border-radius:10px;padding:14px;text-align:center;'>
              <div style='font-size:1.6rem;'>{emoji}</div>
              <div style='font-weight:700;color:{color};font-size:1rem;
                          margin:4px 0;'>Grade {i}</div>
              <div style='font-size:1.5rem;font-weight:800;
                          color:#1e293b;'>{cnt}장</div>
              <div style='font-size:.78rem;color:#64748b;'>{pct:.1f}%</div>
            </div>""", unsafe_allow_html=True)
    st.markdown("")


def grade_badge(grade, confidence):
    info = GRADE_INFO[grade]
    st.markdown(f"""
    <div style='background:{info["color"]}22;border:2px solid {info["color"]};
                border-radius:12px;padding:20px;text-align:center;margin:12px 0;'>
      <div style='font-size:2.8rem;'>{info["emoji"]}</div>
      <div style='font-size:1.6rem;font-weight:700;color:{info["color"]};'>
        {info["label"]}</div>
      <div style='color:#64748b;margin-top:6px;'>
        신뢰도: <b>{confidence:.1f}%</b></div>
    </div>""", unsafe_allow_html=True)


def prob_bars(all_probs):
    st.markdown("**📊 등급별 예측 확률**")
    colors = ["#2ecc71","#f1c40f","#e67e22","#e74c3c","#8e44ad"]
    for i in range(1, 6):
        p = all_probs[i]
        st.markdown(f"""
        <div style='margin:5px 0;'>
          <div style='display:flex;align-items:center;gap:8px;'>
            <span style='width:58px;font-size:.84rem;color:#64748b;'>Grade {i}</span>
            <div style='flex:1;background:#f1f5f9;border-radius:6px;
                        height:18px;overflow:hidden;'>
              <div style='width:{p:.1f}%;background:{colors[i-1]};
                          height:100%;border-radius:6px;'></div>
            </div>
            <span style='width:52px;text-align:right;font-size:.84rem;
                         font-weight:600;color:#334155;'>{p:.1f}%</span>
          </div>
        </div>""", unsafe_allow_html=True)


def show_no_model():
    st.warning("⚠️ **학습된 모델(`model/best_model.pth`)이 없습니다.**\n\n"
               "먼저 `python train.py` 를 실행해 주세요.")
    with st.expander("📋 데이터 준비 방법"):
        st.code("""# labels.csv 형식
filename,grade
IMG_001.jpg,1
IMG_002.jpg,3
""", language="csv")
        st.code("""# 실행 순서
pip install -r requirements.txt
python prepare_labels.py
python train.py
streamlit run app.py
""", language="bash")


def clipboard_tab(model, model_ready):
    st.markdown("#### 📋 캡쳐 후 Ctrl+V 로 바로 붙여넣기")
    st.info("💡 **사용법**: 캡쳐도구(Win+Shift+S) 또는 Print Screen 으로 복사 → "
            "아래 영역 클릭 → **Ctrl+V**")

    if "clipboard_img" not in st.session_state:
        st.session_state.clipboard_img = None

    st.components.v1.html(CLIPBOARD_HTML, height=280)
    st.markdown("---")
    st.markdown("##### 또는 직접 파일로 저장 후 업로드")
    fallback = st.file_uploader(
        "파일 업로드 (.jpg / .png)",
        type=["jpg","jpeg","png"],
        key="clipboard_fallback"
    )

    if fallback:
        img = Image.open(fallback).convert("RGB")
        col1, col2 = st.columns([1, 1], gap="large")
        with col1:
            st.image(img, caption=fallback.name, use_container_width=True)
        with col2:
            if model_ready:
                with st.spinner("🔍 분석 중..."):
                    grade, conf, all_probs = predict(model, img)
                grade_badge(grade, conf)
                prob_bars(all_probs)
            else:
                st.info("모델을 먼저 학습시켜 주세요.")


# ── 메인 ─────────────────────────────────────────
def main():
    render_header()

    model_ready = os.path.exists(MODEL_PATH)
    use_cbam    = False
    if model_ready:
        try:
            model, use_cbam = load_model()
            cbam_txt = " | 🧠 CBAM Attention" if use_cbam else ""
            st.success(f"✅ 모델 로드 완료 | Device: **{DEVICE}**{cbam_txt}")
            render_model_stats()
        except Exception as e:
            st.error(f"❌ 모델 로드 실패: {e}")
            model_ready = False
            model = None
    else:
        show_no_model()
        model = None

    st.markdown("---")

    tab1, tab2, tab3, tab4 = st.tabs([
        "📷 단일 이미지 분석",
        "📋 클립보드 붙여넣기 (Ctrl+V)",
        "📦 일괄 처리 & Excel 추출 (사진 포함)",
        "🗺️ Contact Point 위치 분석 (Grad-CAM)"
    ])

    # ══ TAB 1: 단일 이미지 (+ ROI 슬라이더 지정) ══
    with tab1:
        st.markdown("#### 이미지 1장 업로드 → 등급 즉시 확인")

        # ROI 모드 토글
        use_roi = st.toggle(
            "✏️ Contact Point 직접 지정 (슬라이더 ROI)",
            value=False,
            help="켜면 슬라이더로 배관-Support 접촉 부위를 지정하고 해당 영역만 분석합니다. "
                 "정확도가 향상됩니다."
        )

        uploaded = st.file_uploader(
            "이미지 선택 (.jpg / .jpeg / .png)",
            type=["jpg","jpeg","png"], key="single")

        if uploaded:
            img   = Image.open(uploaded).convert("RGB")
            img_w, img_h = img.size

            if use_roi:
                # ── ROI 슬라이더 모드 ──
                st.markdown("""
                <div style='background:#1e293b;border-radius:10px;
                padding:12px 16px;margin:8px 0;
                border-left:4px solid #f59e0b;'>
                <b style='color:#f59e0b;'>📐 Contact Point 영역 지정</b>
                <span style='color:#94a3b8;font-size:0.87rem;'>
                &nbsp;슬라이더로 주황 박스를 배관-Support 접촉 부위에 맞추세요.
                </span></div>""", unsafe_allow_html=True)

                col_sl, col_prev = st.columns([1, 2], gap="large")

                with col_sl:
                    st.markdown("**가로 범위 (좌% → 우%)**")
                    x_rng = st.slider("좌우", 0, 100, (20, 80),
                                      key=f"xrng_{uploaded.name}",
                                      label_visibility="collapsed")
                    st.markdown("**세로 범위 (위% → 아래%)**")
                    y_rng = st.slider("상하", 0, 100, (50, 90),
                                      key=f"yrng_{uploaded.name}",
                                      label_visibility="collapsed")

                    px1 = int(img_w * x_rng[0] / 100)
                    py1 = int(img_h * y_rng[0] / 100)
                    px2 = int(img_w * x_rng[1] / 100)
                    py2 = int(img_h * y_rng[1] / 100)

                    st.markdown(f"""
                    <div style='background:#0f172a;border-radius:8px;
                    padding:10px;border-left:3px solid #f59e0b;margin-top:8px;'>
                    <span style='color:#f59e0b;font-weight:700;'>박스 좌표</span><br>
                    <span style='color:#e2e8f0;font-size:0.86rem;'>
                    ({px1}, {py1}) → ({px2}, {py2})<br>
                    크기: {px2-px1} × {py2-py1} px
                    </span></div>""", unsafe_allow_html=True)

                    analyze_btn = st.button(
                        "🚀 ROI 적용 분석",
                        type="primary", use_container_width=True,
                        key=f"roi_btn_{uploaded.name}")

                with col_prev:
                    # 박스 미리보기
                    prev = img.copy()
                    drw  = ImageDraw.Draw(prev)
                    lw   = max(3, img_w // 150)
                    drw.rectangle([px1, py1, px2, py2],
                                  outline="#ff6400", width=lw)
                    # 모서리 강조
                    cs = max(12, img_w // 50)
                    for cx_, cy_ in [(px1,py1),(px2,py1),(px1,py2),(px2,py2)]:
                        drw.rectangle([cx_-cs, cy_-cs, cx_+cs, cy_+cs],
                                      outline="#ff6400", width=lw+2)
                    st.image(prev, caption="주황 박스 = 지정된 Contact Point 영역",
                             use_container_width=True)

                if analyze_btn:
                    roi_img  = img
                    box_ok   = (px2-px1) > 20 and (py2-py1) > 20
                    if box_ok:
                        pw = int((px2-px1)*0.08)
                        ph = int((py2-py1)*0.08)
                        roi_img = img.crop((
                            max(0, px1-pw), max(0, py1-ph),
                            min(img_w, px2+pw), min(img_h, py2+ph)
                        ))

                    col_r1, col_r2 = st.columns([1,1], gap="large")
                    with col_r1:
                        st.markdown("**✂️ 잘라낸 ROI 영역**")
                        st.image(roi_img, use_container_width=True)
                    with col_r2:
                        if model_ready:
                            with st.spinner("🔍 분석 중..."):
                                grade, conf, all_probs = predict(model, roi_img)
                            grade_badge(grade, conf)
                            prob_bars(all_probs)
                            if box_ok:
                                st.success("✅ Contact Point ROI 영역으로 분석 완료")
                        else:
                            st.info("모델을 먼저 학습시켜 주세요.")

            else:
                # ── 일반 모드 ──
                col1, col2 = st.columns([1, 1], gap="large")
                with col1:
                    st.image(img, caption=uploaded.name, use_container_width=True)
                with col2:
                    if model_ready:
                        with st.spinner("🔍 분석 중..."):
                            grade, conf, all_probs = predict(model, img)
                        grade_badge(grade, conf)
                        prob_bars(all_probs)
                    else:
                        st.info("모델을 먼저 학습시켜 주세요.")

    # ══ TAB 2: 클립보드 ══
    with tab2:
        clipboard_tab(model, model_ready)

    # ══ TAB 4: Contact Point 위치 분석 ══
    with tab4:
        st.markdown("""<div style='background:linear-gradient(135deg,#0f172a,#1e293b);
        padding:18px 22px;border-radius:12px;margin-bottom:16px;'>
        <h3 style='color:#60a5fa;margin:0;'>🗺️ Contact Point 위치 분석</h3>
        <p style='color:#94a3b8;margin:6px 0 0;font-size:0.9rem;'>
        모델이 실제로 <b style='color:#f59e0b;'>배관-Support 접촉 부위</b>를 보고 판단하는지 시각적으로 확인합니다.
        </p></div>""", unsafe_allow_html=True)

        if not GRADCAM_AVAILABLE:
            st.error("❌ gradcam.py 파일이 없습니다. 같은 폴더에 gradcam.py를 확인하세요.")
        elif not model_ready:
            st.warning("⚠️ 모델을 먼저 학습시켜 주세요 (`python train.py`).")
        else:
            # ── 설명 ──
            with st.expander("📖 히트맵 읽는 법", expanded=True):
                col_l, col_r = st.columns(2)
                with col_l:
                    st.markdown("""
                    **🔥 Grad-CAM (왼쪽)**
                    - 모델이 **등급 판단 시 어느 픽셀을 봤는가**
                    - 🔴 빨강/노랑 = 집중한 영역
                    - 🔵 파랑 = 거의 보지 않은 영역
                    - Contact Point가 빨갛게 나오면 ✅ 올바른 학습
                    """)
                with col_r:
                    st.markdown("""
                    **🧠 CBAM Attention (오른쪽)**
                    - CBAM이 **공간적으로 어디에 집중하는가**
                    - 밝을수록 CBAM이 중요하게 여기는 영역
                    - 배관 접촉부가 밝게 나오면 ✅ CBAM 정상 작동
                    """)

            st.markdown("---")

            # ── 단일 이미지 분석 ──
            st.markdown("#### 📷 이미지 업로드 → Contact Point 히트맵")
            cp_file = st.file_uploader(
                "분석할 이미지 선택",
                type=["jpg","jpeg","png"],
                key="gradcam_single"
            )

            if cp_file:
                pil_img = Image.open(cp_file).convert("RGB")

                with st.spinner("🔬 Grad-CAM 분석 중... (10~20초 소요)"):
                    try:
                        grade, conf, all_probs = predict(model, pil_img)
                        grade_color = GRADE_INFO[grade]["color"]
                        gradcam_ov, cbam_ov, comparison = analyze_contact_point(
                            model, pil_img, IMG_SIZE, DEVICE,
                            grade=grade, grade_color=grade_color)

                        # ── 등급 배지 ──
                        grade_badge(grade, conf)

                        # ── 3분할 비교 이미지 ──
                        st.markdown("##### 📊 Contact Point 히트맵 분석 결과")
                        st.image(comparison, use_container_width=True)

                        # ── 해석 가이드 ──
                        focus_score = float(np.max(
                            np.array(comparison).mean(axis=2)
                        ))
                        st.markdown("---")
                        st.markdown("##### 🔍 판단 가이드")
                        c1, c2 = st.columns(2)
                        with c1:
                            st.markdown("""
                            <div style='background:#1e293b;border-radius:10px;
                            padding:14px;border-left:4px solid #22c55e;'>
                            <b style='color:#22c55e;'>✅ 올바른 학습 패턴</b><br>
                            <span style='color:#94a3b8;font-size:0.87rem;'>
                            배관과 Support/Rack의 접촉 부위(하단/측면)가
                            빨간색으로 강조되는 경우
                            </span></div>""", unsafe_allow_html=True)
                        with c2:
                            st.markdown("""
                            <div style='background:#1e293b;border-radius:10px;
                            padding:14px;border-left:4px solid #ef4444;'>
                            <b style='color:#ef4444;'>⚠️ 재학습 필요 패턴</b><br>
                            <span style='color:#94a3b8;font-size:0.87rem;'>
                            배경, 하늘, 배관 몸통 등
                            Contact Point와 무관한 영역이 강조되는 경우
                            </span></div>""", unsafe_allow_html=True)

                        # ── 다운로드 버튼 ──
                        st.markdown("---")
                        dl_col1, dl_col2 = st.columns(2)
                        with dl_col1:
                            st.download_button(
                                "💾 Grad-CAM 이미지 저장",
                                data=pil_to_bytes(gradcam_ov),
                                file_name=f"gradcam_{cp_file.name}",
                                mime="image/png",
                                use_container_width=True
                            )
                        with dl_col2:
                            st.download_button(
                                "💾 CBAM Attention 이미지 저장",
                                data=pil_to_bytes(cbam_ov),
                                file_name=f"cbam_{cp_file.name}",
                                mime="image/png",
                                use_container_width=True
                            )

                    except Exception as e:
                        st.error(f"❌ 분석 오류: {e}")
                        import traceback
                        st.code(traceback.format_exc())

            st.markdown("---")

            # ── 일괄 Contact Point 분석 ──
            st.markdown("#### 📦 일괄 업로드 → 전체 히트맵 생성")
            st.info("💡 여러 장을 한 번에 분석하여 모델이 어디를 보는지 패턴을 파악할 수 있습니다.")

            batch_cp_files = st.file_uploader(
                "이미지 여러 장 선택 (최대 10장)",
                type=["jpg","jpeg","png"],
                accept_multiple_files=True,
                key="gradcam_batch"
            )

            if batch_cp_files:
                if len(batch_cp_files) > 10:
                    st.warning("⚠️ 최대 10장까지 지원합니다. 처음 10장만 처리합니다.")
                    batch_cp_files = batch_cp_files[:10]

                if st.button("🔬 일괄 Contact Point 분석",
                             type="primary", use_container_width=True,
                             key="gradcam_batch_btn"):
                    prog = st.progress(0, text="히트맵 생성 중...")
                    for bi, bf in enumerate(batch_cp_files):
                        pil_b = Image.open(bf).convert("RGB")
                        try:
                            g, c, _ = predict(model, pil_b)
                            gc = GRADE_INFO[g]["color"]
                            _, _, comp_b = analyze_contact_point(
                                model, pil_b, IMG_SIZE, DEVICE,
                                grade=g, grade_color=gc)

                            st.markdown(f"**{bf.name}** — "
                                        f"{GRADE_INFO[g]['emoji']} Grade {g} "
                                        f"(신뢰도 {c:.1f}%)")
                            st.image(comp_b, use_container_width=True)
                            st.markdown("")
                        except Exception as e:
                            st.error(f"{bf.name}: 오류 — {e}")
                        prog.progress((bi+1)/len(batch_cp_files),
                                      text=f"처리 중 {bi+1}/{len(batch_cp_files)}")
                    prog.empty()
                    st.success(f"✅ {len(batch_cp_files)}장 분석 완료")

    # ══ TAB 3: 일괄 처리 + Excel(사진 포함) ══
    with tab3:
        st.markdown("#### 여러 장 업로드 → 전체 분석 → Excel 다운로드 (사진 포함)")
        files = st.file_uploader(
            "이미지 여러 장 선택 (.jpg / .jpeg / .png)",
            type=["jpg","jpeg","png"],
            accept_multiple_files=True, key="batch")

        if files:
            st.info(f"📂 **{len(files)}장** 이미지 업로드됨")

            if not model_ready:
                st.warning("모델을 먼저 학습시켜 주세요.")
            else:
                if st.button("🚀 일괄 분석 시작", type="primary",
                             use_container_width=True):
                    results   = []
                    pil_imgs  = []      # Excel 사진 포함용
                    progress  = st.progress(0, text="분석 중...")
                    n_prev    = min(5, len(files))
                    prev_cols = st.columns(n_prev)

                    for i, f in enumerate(files):
                        img   = Image.open(f).convert("RGB")
                        grade, conf, _ = predict(model, img)
                        info  = GRADE_INFO[grade]
                        results.append({
                            "파일명"    : f.name,
                            "예측 등급" : grade,
                            "등급 설명" : info["label"],
                            "신뢰도(%)": round(conf, 1),
                            "분석 시각" : datetime.datetime.now()
                                         .strftime("%Y-%m-%d %H:%M:%S"),
                        })
                        pil_imgs.append(img)

                        if i < 5:
                            with prev_cols[i]:
                                st.image(img, caption=f.name,
                                         use_container_width=True)
                                st.markdown(
                                    f"<div style='text-align:center;"
                                    f"color:{info['color']};font-weight:700;'>"
                                    f"{info['emoji']} Grade {grade}</div>",
                                    unsafe_allow_html=True)
                        progress.progress((i+1)/len(files),
                                          text=f"분석 중... {i+1}/{len(files)}")

                    progress.empty()
                    st.success(f"✅ 분석 완료! 총 **{len(results)}장** 처리됨")

                    df = pd.DataFrame(results)
                    c1, c2, c3 = st.columns(3)
                    c1.metric("총 이미지",      f"{len(df)}장")
                    c2.metric("평균 등급",       f"{df['예측 등급'].mean():.1f}")
                    high = len(df[df["예측 등급"] >= 4])
                    c3.metric("고위험(4~5등급)", f"{high}장",
                              delta=f"{high/len(df)*100:.0f}%",
                              delta_color="inverse")

                    st.markdown("**📊 등급 분포**")
                    st.bar_chart(
                        df["예측 등급"].value_counts().sort_index(),
                        color="#3b82f6")

                    st.dataframe(
                        df, use_container_width=True, hide_index=True,
                        column_config={
                            "예측 등급": st.column_config.NumberColumn(
                                "예측 등급", format="%d"),
                            "신뢰도(%)": st.column_config.ProgressColumn(
                                "신뢰도(%)", min_value=0,
                                max_value=100, format="%.1f%%"),
                        }
                    )

                    st.markdown("---")
                    st.markdown("#### 📥 Excel 다운로드 옵션")
                    col_a, col_b = st.columns(2)

                    fname_base = ("corrosion_result_"
                                  + datetime.datetime.now()
                                  .strftime("%Y%m%d_%H%M%S"))
                    mime = ("application/vnd.openxmlformats-officedocument"
                            ".spreadsheetml.sheet")

                    with col_a:
                        st.download_button(
                            label="📸 Excel 다운로드 (사진 포함)",
                            data=to_excel_with_images(df, pil_imgs),
                            file_name=fname_base + "_with_photos.xlsx",
                            mime=mime,
                            use_container_width=True,
                            type="primary",
                            help="각 행에 분석 사진 썸네일이 포함된 Excel"
                        )
                    with col_b:
                        # 사진 없는 간단 버전도 유지
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine="openpyxl") as writer:
                            df.to_excel(writer, index=False, sheet_name="부식등급결과")
                        st.download_button(
                            label="📄 Excel 다운로드 (데이터만)",
                            data=output.getvalue(),
                            file_name=fname_base + "_data_only.xlsx",
                            mime=mime,
                            use_container_width=True,
                            type="secondary",
                            help="사진 없이 데이터만 포함된 가벼운 Excel"
                        )

    st.markdown("---")
    st.markdown(
        "<div style='text-align:center;color:#94a3b8;font-size:.8rem;'>"
        "🏭 Contact Point Corrosion Grade Classifier &nbsp;|&nbsp; "
        "EfficientNet-B0 + CBAM Attention &nbsp;|&nbsp; "
        "PyTorch + Streamlit</div>",
        unsafe_allow_html=True)


if __name__ == "__main__":
    main()
