# ============================================================
#  app.py  (ROI-aligned inference version)
#  - train.py 의 ROI crop 학습 방식과 추론 일관성 맞춤
#  - ROI detector 있으면 자동 ROI
#  - 수동 ROI 슬라이더로 보정 가능
#  - 단일 이미지 / 배치 처리 / Excel 저장 / Grad-CAM 지원
# ============================================================

import os
import io
import json
import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image, ImageDraw

import torch
import torch.nn as nn
from torchvision import models, transforms

# ────────────────────────────────────────────────
# 기본 설정
# ────────────────────────────────────────────────
st.set_page_config(
    page_title="Contact Point Corrosion 등급 분류",
    page_icon="🔬",
    layout="wide",
)

BASE_DIR = Path(".")
MODEL_PATH = BASE_DIR / "model" / "best_model.pth"
ROI_MODEL_PATH = BASE_DIR / "model" / "roi_detector.pth"
ROI_JSON_PATH = BASE_DIR / "data" / "roi_annotations.json"

DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")
DEFAULT_IMG_SIZE = 224
DEFAULT_NUM_CLASSES = 5

GRADE_INFO = {
    1: {"label": "Grade 1 — 정상",         "color": "#2ecc71", "emoji": "🟢"},
    2: {"label": "Grade 2 — 경미한 스케일", "color": "#f1c40f", "emoji": "🟡"},
    3: {"label": "Grade 3 — 중간 스케일",   "color": "#e67e22", "emoji": "🟠"},
    4: {"label": "Grade 4 — 심각한 스케일", "color": "#e74c3c", "emoji": "🔴"},
    5: {"label": "Grade 5 — 매우 심각",     "color": "#8e44ad", "emoji": "🟣"},
}

# Grad-CAM
try:
    from gradcam import analyze_contact_point, pil_to_bytes
    GRADCAM_AVAILABLE = True
except Exception:
    GRADCAM_AVAILABLE = False


# ────────────────────────────────────────────────
# 모델 정의 (train.py 와 동일 계열)
# ────────────────────────────────────────────────
class ChannelAttention(nn.Module):
    def __init__(self, in_channels, reduction=16):
        super().__init__()
        r = max(1, in_channels // reduction)
        self.avg_pool = nn.AdaptiveAvgPool2d(1)
        self.max_pool = nn.AdaptiveMaxPool2d(1)
        self.fc = nn.Sequential(
            nn.Linear(in_channels, r, bias=False),
            nn.ReLU(inplace=True),
            nn.Linear(r, in_channels, bias=False),
        )
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        b, c, _, _ = x.size()
        avg = self.fc(self.avg_pool(x).view(b, c))
        mx = self.fc(self.max_pool(x).view(b, c))
        attn = self.sigmoid(avg + mx).view(b, c, 1, 1)
        return x * attn


class SpatialAttention(nn.Module):
    def __init__(self, kernel_size=7):
        super().__init__()
        self.conv = nn.Conv2d(2, 1, kernel_size, padding=kernel_size // 2, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        avg = x.mean(dim=1, keepdim=True)
        mx, _ = x.max(dim=1, keepdim=True)
        attn = self.sigmoid(self.conv(torch.cat([avg, mx], dim=1)))
        return x * attn


class CBAM(nn.Module):
    def __init__(self, in_channels, reduction=16, kernel_size=7):
        super().__init__()
        self.ch = ChannelAttention(in_channels, reduction)
        self.sp = SpatialAttention(kernel_size)

    def forward(self, x):
        return self.sp(self.ch(x))


class EfficientNetWithCBAM(nn.Module):
    def __init__(self, num_classes):
        super().__init__()
        base = models.efficientnet_b0(weights=None)
        self.features = base.features
        self.cbam = CBAM(1280)
        self.pool = nn.AdaptiveAvgPool2d(1)
        self.classifier = nn.Sequential(
            nn.Dropout(p=0.5),
            nn.Linear(1280, 256),
            nn.ReLU(inplace=True),
            nn.Dropout(p=0.3),
            nn.Linear(256, num_classes),
        )

    def forward(self, x):
        feat = self.features(x)
        feat = self.cbam(feat)
        feat = self.pool(feat).flatten(1)
        return self.classifier(feat)


class ROIRegressor(nn.Module):
    def __init__(self):
        super().__init__()
        base = models.mobilenet_v3_small(weights=None)
        in_f = base.classifier[-1].in_features
        base.classifier[-1] = nn.Linear(in_f, 4)
        self.model = base

    def forward(self, x):
        return self.model(x)


# ────────────────────────────────────────────────
# 로딩
# ────────────────────────────────────────────────
@st.cache_resource
def load_classifier():
    if not MODEL_PATH.exists():
        return None, None

    ckpt = torch.load(MODEL_PATH, map_location=DEVICE)
    num_classes = ckpt.get("num_classes", DEFAULT_NUM_CLASSES)
    use_cbam = ckpt.get("use_cbam", True)

    if use_cbam:
        model = EfficientNetWithCBAM(num_classes)
    else:
        base = models.efficientnet_b0(weights=None)
        in_f = base.classifier[1].in_features
        base.classifier = nn.Sequential(
            nn.Dropout(p=0.5),
            nn.Linear(in_f, 256),
            nn.ReLU(inplace=True),
            nn.Dropout(p=0.3),
            nn.Linear(256, num_classes),
        )
        model = base

    model.load_state_dict(ckpt["model_state_dict"])
    model.to(DEVICE).eval()
    return model, ckpt


@st.cache_resource
def load_roi_detector():
    if not ROI_MODEL_PATH.exists():
        return None

    try:
        ckpt = torch.load(ROI_MODEL_PATH, map_location=DEVICE)
        model = ROIRegressor()
        if isinstance(ckpt, dict) and "model_state_dict" in ckpt:
            model.load_state_dict(ckpt["model_state_dict"])
        else:
            model.load_state_dict(ckpt)
        model.to(DEVICE).eval()
        return model
    except Exception:
        return None


@st.cache_data
def load_roi_json():
    if ROI_JSON_PATH.exists():
        try:
            with open(ROI_JSON_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


# ────────────────────────────────────────────────
# 유틸
# ────────────────────────────────────────────────
def get_model_info(ckpt):
    if ckpt is None:
        return None
    return {
        "total_data": ckpt.get("total_data", "?"),
        "total_train": ckpt.get("total_train", "?"),
        "best_val_accuracy": ckpt.get("best_val_accuracy", 0.0),
        "grade_dist": ckpt.get("grade_dist", {}),
        "use_cbam": ckpt.get("use_cbam", False),
        "use_roi_crop": ckpt.get("use_roi_crop", False),
        "roi_expand_ratio": ckpt.get("roi_expand_ratio", 0.15),
        "img_size": ckpt.get("img_size", DEFAULT_IMG_SIZE),
    }


def get_classifier_transform(img_size):
    return transforms.Compose([
        transforms.Resize((img_size, img_size)),
        transforms.ToTensor(),
        transforms.Normalize([0.485, 0.456, 0.406],
                             [0.229, 0.224, 0.225]),
    ])


def get_roi_transform():
    return transforms.Compose([
        transforms.Resize((224, 224)),
        transforms.ToTensor(),
        transforms.Normalize([0.485, 0.456, 0.406],
                             [0.229, 0.224, 0.225]),
    ])


def clamp_box(x1, y1, x2, y2, w, h):
    x1 = int(max(0, min(x1, w - 1)))
    y1 = int(max(0, min(y1, h - 1)))
    x2 = int(max(1, min(x2, w)))
    y2 = int(max(1, min(y2, h)))
    if x2 <= x1:
        x2 = min(w, x1 + 1)
    if y2 <= y1:
        y2 = min(h, y1 + 1)
    return x1, y1, x2, y2


def expand_box(x1, y1, x2, y2, w, h, expand_ratio=0.15):
    bw = x2 - x1
    bh = y2 - y1
    pad_x = int(bw * expand_ratio)
    pad_y = int(bh * expand_ratio)
    return clamp_box(x1 - pad_x, y1 - pad_y, x2 + pad_x, y2 + pad_y, w, h)


def crop_image_with_box(img, box, expand_ratio=0.15):
    w, h = img.size
    x1, y1, x2, y2 = box
    x1, y1, x2, y2 = expand_box(x1, y1, x2, y2, w, h, expand_ratio)
    return img.crop((x1, y1, x2, y2)), (x1, y1, x2, y2)


def draw_box(img, box, color="#ff6400", width=4):
    out = img.copy()
    draw = ImageDraw.Draw(out)
    x1, y1, x2, y2 = box
    draw.rectangle([x1, y1, x2, y2], outline=color, width=width)
    return out


def pil_to_jpg_bytes(pil_img, quality=92):
    buf = io.BytesIO()
    pil_img.convert("RGB").save(buf, format="JPEG", quality=quality)
    return buf.getvalue()


def predict_grade(model, pil_img, img_size):
    tf = get_classifier_transform(img_size)
    x = tf(pil_img).unsqueeze(0).to(DEVICE)
    with torch.no_grad():
        probs = torch.softmax(model(x), dim=1).squeeze(0)
    pred_grade = int(torch.argmax(probs).item()) + 1
    conf = float(torch.max(probs).item()) * 100.0
    all_probs = {i + 1: float(probs[i].item()) * 100.0 for i in range(len(probs))}
    return pred_grade, conf, all_probs


def predict_roi_box(roi_model, pil_img):
    if roi_model is None:
        return None

    w, h = pil_img.size
    tf = get_roi_transform()
    x = tf(pil_img).unsqueeze(0).to(DEVICE)

    with torch.no_grad():
        pred = roi_model(x).squeeze(0).cpu().numpy()

    if np.max(pred) <= 2.0:
        x1, y1, x2, y2 = pred
        x1 *= w
        x2 *= w
        y1 *= h
        y2 *= h
    else:
        x1, y1, x2, y2 = pred

    return clamp_box(x1, y1, x2, y2, w, h)


def get_roi_from_json(filename, roi_map):
    roi = roi_map.get(filename)
    if not roi:
        return None

    try:
        x1 = int(round(float(roi["x1"])))
        y1 = int(round(float(roi["y1"])))
        x2 = int(round(float(roi["x2"])))
        y2 = int(round(float(roi["y2"])))
        return x1, y1, x2, y2
    except Exception:
        return None


def grade_description(grade):
    return GRADE_INFO[grade]["label"]


# ────────────────────────────────────────────────
# Excel 저장
# ────────────────────────────────────────────────
def to_excel_with_images(df: pd.DataFrame, images: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = "부식등급결과"

    headers = ["사진"] + list(df.columns)
    ws.append(headers)

    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    THUMB_W = 80
    THUMB_H = 80
    ROW_H = 65

    grade_colors = {
        "1": "C6EFCE", "2": "FFEB9C", "3": "FFCC99", "4": "FFC7CE", "5": "E2AFEF"
    }

    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        ws.row_dimensions[row_idx].height = ROW_H
        ws.cell(row=row_idx, column=1).value = ""

        for col_idx, val in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if df.columns[col_idx - 2] == "예측 등급":
                gs = str(val)
                if gs in grade_colors:
                    cell.fill = PatternFill(
                        start_color=grade_colors[gs],
                        end_color=grade_colors[gs],
                        fill_type="solid"
                    )

        if row_idx - 2 < len(images) and images[row_idx - 2] is not None:
            try:
                pil_img = images[row_idx - 2].copy()
                pil_img.thumbnail((THUMB_W, THUMB_H))
                img_buf = io.BytesIO()
                pil_img.save(img_buf, format="PNG")
                img_buf.seek(0)
                xl_img = XLImage(img_buf)
                xl_img.width = THUMB_W
                xl_img.height = THUMB_H
                ws.add_image(xl_img, f"A{row_idx}")
            except Exception:
                pass

    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ────────────────────────────────────────────────
# UI
# ────────────────────────────────────────────────
def render_header():
    st.markdown("""
    <div style='background:linear-gradient(135deg,#1a1a2e,#16213e);
                padding:28px 32px;border-radius:14px;margin-bottom:20px;'>
      <h1 style='color:#e2e8f0;margin:0;font-size:1.9rem;'>
        🔬 Contact Point Corrosion 등급 분류
      </h1>
      <p style='color:#94a3b8;margin:8px 0 0;'>
        ROI Crop 학습/추론 일관성 적용 버전
      </p>
    </div>
    """, unsafe_allow_html=True)


def render_model_stats(info):
    if not info:
        return

    acc = info["best_val_accuracy"] * 100
    cbam_badge = "🧠 CBAM 적용" if info.get("use_cbam") else "기본 EfficientNet"
    roi_badge = "🎯 ROI Crop 사용" if info.get("use_roi_crop") else "원본 전체 이미지"

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📸 전체 학습 데이터", f"{info['total_data']}장")
    c2.metric("🎓 학습 사용 데이터", f"{info['total_train']}장")
    c3.metric("🎯 검증 정확도", f"{acc:.1f}%")
    c4.metric("⚙️ 추론 방식", "ROI Crop" if info.get("use_roi_crop") else "Full Image")

    st.info(f"{cbam_badge} | {roi_badge} | ROI expand ratio = {info.get('roi_expand_ratio', 0.15):.2f}")

    grade_dist = info.get("grade_dist", {})
    if grade_dist:
        st.markdown("#### 📊 등급별 학습 데이터 현황")
        cols = st.columns(5)
        total = sum(grade_dist.values()) if isinstance(grade_dist, dict) else 0
        for i in range(1, 6):
            cnt = grade_dist.get(i, grade_dist.get(str(i), 0)) if isinstance(grade_dist, dict) else 0
            pct = (cnt / total * 100) if total > 0 else 0
            g = GRADE_INFO[i]
            cols[i - 1].markdown(f"""
            <div style='background:{g["color"]}18;border:1.5px solid {g["color"]};
                        border-radius:10px;padding:14px;text-align:center;'>
              <div style='font-size:1.5rem;'>{g["emoji"]}</div>
              <div style='font-weight:700;color:{g["color"]};'>Grade {i}</div>
              <div style='font-size:1.5rem;font-weight:800;color:#1e293b;'>{cnt}장</div>
              <div style='font-size:.8rem;color:#64748b;'>{pct:.1f}%</div>
            </div>
            """, unsafe_allow_html=True)


def grade_badge(grade, confidence):
    info = GRADE_INFO[grade]
    st.markdown(f"""
    <div style='background:{info["color"]}22;border:2px solid {info["color"]};
                border-radius:12px;padding:20px;text-align:center;margin:12px 0;'>
      <div style='font-size:2.8rem;'>{info["emoji"]}</div>
      <div style='font-size:1.6rem;font-weight:700;color:{info["color"]};'>
        {info["label"]}</div>
      <div style='color:#64748b;margin-top:6px;'>신뢰도: <b>{confidence:.1f}%</b></div>
    </div>
    """, unsafe_allow_html=True)


def prob_bars(all_probs):
    st.markdown("**📊 등급별 예측 확률**")
    colors = ["#2ecc71", "#f1c40f", "#e67e22", "#e74c3c", "#8e44ad"]
    for i in range(1, 6):
        p = all_probs.get(i, 0.0)
        st.markdown(f"""
        <div style='margin:5px 0;'>
          <div style='display:flex;align-items:center;gap:8px;'>
            <span style='width:58px;font-size:.84rem;color:#64748b;'>Grade {i}</span>
            <div style='flex:1;background:#f1f5f9;border-radius:6px;height:18px;overflow:hidden;'>
              <div style='width:{p:.1f}%;background:{colors[i-1]};height:100%;border-radius:6px;'></div>
            </div>
            <span style='width:52px;text-align:right;font-size:.84rem;font-weight:600;color:#334155;'>{p:.1f}%</span>
          </div>
        </div>
        """, unsafe_allow_html=True)


def show_no_model():
    st.warning("⚠️ `model/best_model.pth` 가 없습니다. 먼저 `python train.py` 를 실행하세요.")


def analyze_single_image(
    model,
    info,
    roi_model,
    roi_map,
    pil_img,
    filename_for_json=None,
    use_manual_roi=False,
    manual_box=None,
    use_auto_roi=True,
):
    img_size = info.get("img_size", DEFAULT_IMG_SIZE) if info else DEFAULT_IMG_SIZE
    roi_expand_ratio = info.get("roi_expand_ratio", 0.15) if info else 0.15
    use_roi_crop = info.get("use_roi_crop", False) if info else False

    original = pil_img.copy()
    used_mode = "원본 전체"

    box = None

    if use_manual_roi and manual_box is not None:
        box = manual_box
        used_mode = "수동 ROI"
    elif use_auto_roi:
        if filename_for_json:
            box = get_roi_from_json(filename_for_json, roi_map)
            if box is not None:
                used_mode = "저장된 ROI"
        if box is None and roi_model is not None:
            box = predict_roi_box(roi_model, original)
            if box is not None:
                used_mode = "자동 ROI"

    roi_preview = None
    infer_img = original

    if use_roi_crop and box is not None:
        roi_preview = draw_box(original, box)
        infer_img, used_box = crop_image_with_box(original, box, roi_expand_ratio)
        box = used_box
    else:
        box = None

    grade, conf, all_probs = predict_grade(model, infer_img, img_size)

    return {
        "grade": grade,
        "confidence": conf,
        "all_probs": all_probs,
        "original_image": original,
        "roi_preview": roi_preview,
        "crop_image": infer_img,
        "used_mode": used_mode,
        "box": box,
        "use_roi_crop": use_roi_crop,
    }


# ────────────────────────────────────────────────
# 메인
# ────────────────────────────────────────────────
def main():
    render_header()

    model, ckpt = load_classifier()
    if model is None:
        show_no_model()
        return

    info = get_model_info(ckpt)
    roi_model = load_roi_detector()
    roi_map = load_roi_json()

    st.success(f"✅ 모델 로드 완료 | Device: **{DEVICE}**")
    render_model_stats(info)

    tab1, tab2, tab3, tab4 = st.tabs([
        "📷 단일 이미지 분석",
        "📦 일괄 처리 & Excel 추출",
        "🗺️ Grad-CAM 분석",
        "ℹ️ ROI 사용 가이드",
    ])

    # ------------------------------------------------
    # TAB 1
    # ------------------------------------------------
    with tab1:
        st.markdown("#### 이미지 1장 업로드 → ROI 기준 등급 분석")

        uploaded = st.file_uploader(
            "이미지 선택 (.jpg / .jpeg / .png)",
            type=["jpg", "jpeg", "png"],
            key="single_img"
        )

        if uploaded:
            img = Image.open(uploaded).convert("RGB")
            w, h = img.size

            c0, c1, c2 = st.columns([1.2, 1.2, 1.2])
            with c0:
                use_auto_roi = st.toggle(
                    "자동 ROI 사용",
                    value=True,
                    help="저장된 ROI 또는 roi_detector.pth 를 이용해 ROI를 자동 추정합니다."
                )
            with c1:
                use_manual_roi = st.toggle(
                    "수동 ROI 사용",
                    value=False,
                    help="켜면 아래 슬라이더 ROI가 자동 ROI보다 우선합니다."
                )
            with c2:
                st.markdown(f"**원본 크기:** {w} × {h}")

            manual_box = None
            if use_manual_roi:
                xr = st.slider("가로 범위 (%)", 0, 100, (20, 80), key="single_x")
                yr = st.slider("세로 범위 (%)", 0, 100, (45, 90), key="single_y")
                x1 = int(w * xr[0] / 100)
                x2 = int(w * xr[1] / 100)
                y1 = int(h * yr[0] / 100)
                y2 = int(h * yr[1] / 100)
                manual_box = clamp_box(x1, y1, x2, y2, w, h)
                preview = draw_box(img, manual_box)
                st.image(preview, caption="수동 ROI 미리보기", use_container_width=True)

            if st.button("🔍 분석 시작", type="primary", use_container_width=True):
                with st.spinner("분석 중..."):
                    result = analyze_single_image(
                        model=model,
                        info=info,
                        roi_model=roi_model,
                        roi_map=roi_map,
                        pil_img=img,
                        filename_for_json=uploaded.name,
                        use_manual_roi=use_manual_roi,
                        manual_box=manual_box,
                        use_auto_roi=use_auto_roi,
                    )

                left, right = st.columns([1.2, 1.0], gap="large")

                with left:
                    if result["roi_preview"] is not None:
                        st.image(result["roi_preview"], caption=f"원본 + ROI 박스 ({result['used_mode']})", use_container_width=True)
                        st.image(result["crop_image"], caption="실제 분류에 사용된 ROI Crop", use_container_width=True)
                    else:
                        st.image(result["original_image"], caption="원본 전체 이미지", use_container_width=True)

                    if result["box"] is not None:
                        x1, y1, x2, y2 = result["box"]
                        st.info(f"ROI 좌표: ({x1}, {y1}) → ({x2}, {y2}) | 방식: {result['used_mode']}")
                    else:
                        st.info(f"ROI 없음 → 원본 전체 사용 | 방식: {result['used_mode']}")

                with right:
                    grade_badge(result["grade"], result["confidence"])
                    prob_bars(result["all_probs"])

                    st.markdown("**분석 요약**")
                    st.write(f"- 분류 방식: {'ROI Crop' if result['use_roi_crop'] else 'Full Image'}")
                    st.write(f"- ROI 선택 방식: {result['used_mode']}")
                    st.write(f"- 예측 등급: Grade {result['grade']}")
                    st.write(f"- 설명: {grade_description(result['grade'])}")

    # ------------------------------------------------
    # TAB 2
    # ------------------------------------------------
    with tab2:
        st.markdown("#### 여러 장 업로드 → ROI 기반 일괄 분석 + Excel 저장")

        files = st.file_uploader(
            "다중 업로드",
            type=["jpg", "jpeg", "png"],
            accept_multiple_files=True,
            key="batch_imgs"
        )

        batch_auto_roi = st.toggle("일괄 처리 시 자동 ROI 사용", value=True, key="batch_auto_roi")

        if files:
            if st.button("📦 일괄 분석 실행", type="primary"):
                rows = []
                excel_images = []

                prog = st.progress(0)
                status = st.empty()

                for i, f in enumerate(files, start=1):
                    status.info(f"처리 중: {f.name} ({i}/{len(files)})")
                    img = Image.open(f).convert("RGB")

                    result = analyze_single_image(
                        model=model,
                        info=info,
                        roi_model=roi_model,
                        roi_map=roi_map,
                        pil_img=img,
                        filename_for_json=f.name,
                        use_manual_roi=False,
                        manual_box=None,
                        use_auto_roi=batch_auto_roi,
                    )

                    rows.append({
                        "파일명": f.name,
                        "예측 등급": result["grade"],
                        "설명": grade_description(result["grade"]),
                        "신뢰도(%)": round(result["confidence"], 2),
                        "ROI 방식": result["used_mode"],
                        "분석 시각": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    })

                    excel_images.append(result["roi_preview"] if result["roi_preview"] is not None else result["original_image"])
                    prog.progress(i / len(files))

                status.success("일괄 분석 완료")
                df = pd.DataFrame(rows)
                st.dataframe(df, use_container_width=True)

                xlsx_bytes = to_excel_with_images(df, excel_images)
                st.download_button(
                    "📥 Excel 다운로드",
                    data=xlsx_bytes,
                    file_name=f"corrosion_batch_result_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # ------------------------------------------------
    # TAB 3
    # ------------------------------------------------
    with tab3:
        st.markdown("#### Contact Point 위치 분석 (Grad-CAM)")
        if not GRADCAM_AVAILABLE:
            st.warning("`gradcam.py` 모듈을 불러오지 못했습니다.")
        else:
            cam_file = st.file_uploader(
                "Grad-CAM 분석 이미지 업로드",
                type=["jpg", "jpeg", "png"],
                key="cam_img"
            )
            if cam_file:
                img = Image.open(cam_file).convert("RGB")
                st.image(img, caption=cam_file.name, use_container_width=True)

                cam_auto_roi = st.toggle(
                    "Grad-CAM에서도 자동 ROI 기준 예측 사용",
                    value=True,
                    key="cam_auto_roi"
                )

                if st.button("🗺️ Grad-CAM 실행", type="primary", key="run_cam"):
                    with st.spinner("Grad-CAM 분석 중..."):
                        try:
                            img_size = info.get("img_size", DEFAULT_IMG_SIZE) if info else DEFAULT_IMG_SIZE

                            pred_result = analyze_single_image(
                                model=model,
                                info=info,
                                roi_model=roi_model,
                                roi_map=roi_map,
                                pil_img=img,
                                filename_for_json=cam_file.name,
                                use_manual_roi=False,
                                manual_box=None,
                                use_auto_roi=cam_auto_roi,
                            )

                            grade = pred_result["grade"]
                            grade_color = GRADE_INFO[grade]["color"]

                            cam_input_img = pred_result["crop_image"] if pred_result["crop_image"] is not None else img

                            gradcam_ov, cbam_ov, comparison_img = analyze_contact_point(
                                model,
                                cam_input_img,
                                img_size,
                                DEVICE,
                                grade=grade,
                                grade_color=grade_color,
                            )

                            st.success(f"예측 등급: {GRADE_INFO[grade]['emoji']} Grade {grade}")
                            st.image(comparison_img, caption="원본 / Grad-CAM / CBAM 비교", use_container_width=True)

                            c1, c2 = st.columns(2)
                            with c1:
                                st.image(gradcam_ov, caption="Grad-CAM Overlay", use_container_width=True)
                                try:
                                    st.download_button(
                                        "💾 Grad-CAM 이미지 저장",
                                        data=pil_to_bytes(gradcam_ov),
                                        file_name=f"gradcam_{cam_file.name}",
                                        mime="image/png",
                                        use_container_width=True,
                                        key="dl_gradcam_single"
                                    )
                                except Exception:
                                    pass

                            with c2:
                                st.image(cbam_ov, caption="CBAM Attention Overlay", use_container_width=True)
                                try:
                                    st.download_button(
                                        "💾 CBAM 이미지 저장",
                                        data=pil_to_bytes(cbam_ov),
                                        file_name=f"cbam_{cam_file.name}",
                                        mime="image/png",
                                        use_container_width=True,
                                        key="dl_cbam_single"
                                    )
                                except Exception:
                                    pass

                            if pred_result["box"] is not None:
                                x1, y1, x2, y2 = pred_result["box"]
                                st.info(
                                    f"Grad-CAM 입력 기준: {pred_result['used_mode']} | "
                                    f"ROI 좌표: ({x1}, {y1}) → ({x2}, {y2})"
                                )
                            else:
                                st.info(f"Grad-CAM 입력 기준: {pred_result['used_mode']}")

                        except Exception as e:
                            st.error(f"Grad-CAM 분석 실패: {e}")
                            import traceback
                            st.code(traceback.format_exc())

    # ------------------------------------------------
    # TAB 4
    # ------------------------------------------------
    with tab4:
        st.markdown("""
        ### ROI 사용 추천 순서

        **1. 저장된 ROI가 있는 파일**
        - `data/roi_annotations.json` 에 있으면 그 ROI를 우선 사용

        **2. 저장된 ROI가 없으면 자동 ROI**
        - `model/roi_detector.pth` 가 있으면 자동으로 Contact Point 추정

        **3. 자동 ROI가 마음에 안 들면 수동 ROI**
        - 단일 이미지 분석 탭에서 수동 ROI 슬라이더로 보정

        ### ROI가 부정확할 때 개선 팁
        - `annotate.py` 에서 틀린 ROI 수동 수정
        - 수정한 ROI를 `roi_annotations.json` 에 저장
        - `train_roi_detector.py` 다시 학습
        - `auto_roi.py --preview` 로 결과 점검
        - 이후 `train.py` 재학습

        ### 지금 app.py 특징
        - 학습 시 ROI crop 사용 여부를 모델 체크포인트에서 자동 반영
        - 추론도 같은 ROI expand ratio 로 crop
        - ROI 없으면 원본 전체 fallback
        """)


if __name__ == "__main__":
    main()